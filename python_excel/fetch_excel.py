# Fetch the data from the excel file using python module openpyxl
# importing the module
import openpyxl

# give the location of your file
path = "data9.xlsx"

# creating an object
data9_workbook = openpyxl.load_workbook(path) # can use "data9.xlsx" instead of path, but path variable is much easier
                                              # if we need to use it a lot of time and the file is not in the same folder

# storing the data from active sheet into an object called data9
data9_sheet = data9_workbook.active

# get the title of the sheet
data9_sheet_title = data9_sheet.title

get_data = data9_sheet.cell(row=1, column=1)

print(get_data.value)
print(data9_sheet_title)

# print total number of rows and columns
count_column = data9_sheet.max_column
count_row = data9_sheet.max_row
print("Number of rows: " + str(count_row) + "; Number of columns: " + str(count_column))

# rename sheet
# go to the python file you create the excel file and rename there

# change value (but not in excel)
data9_sheet.cell(row=1, column=1).value = "changed"
print(get_data.value)

data9_workbook.save(str(path))