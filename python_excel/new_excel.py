# write a program in python to copy data from data9 file to new file called data9_copied

import openpyxl
import xlsxwriter

data9_copied = xlsxwriter.Workbook("data9_copied.xlsx")
worksheet = data9_copied.add_worksheet("Trainees")
data9_copied.close()

# opening the source excel file
path = "data9.xlsx"
data9_workbook = openpyxl.load_workbook(path)
data9_worksheet = data9_workbook.worksheets[0]

# opening the destination excel file
new_path = "data9_copied.xlsx"
data9_copied_workbook = openpyxl.load_workbook(new_path)
data9_copied_worksheet = data9_copied_workbook.active

number_of_column = data9_worksheet.max_column
number_of_row = data9_worksheet.max_row

# copying the cell values from data9 to data9_copied
for i in range(1, number_of_row + 1):
    for j in range(1, number_of_column + 1):
        cell_content = data9_worksheet.cell(row = i, column = j) # fetching cell value from data9

        # writing the read value to data9_copied
        data9_copied_worksheet.cell(row = i, column = j).value = cell_content.value

data9_copied_workbook.save(str(new_path))
